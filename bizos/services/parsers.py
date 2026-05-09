import pandas as pd
import pdfplumber
import openpyxl
import json
from pathlib import Path


def parse_file(file_path: str, file_type: str) -> dict:
    """
    Routes a file to the correct parser based on its type.
    Returns a dictionary of extracted data ready for AI analysis.
    """
    path = Path(file_path)

    if file_type == "csv":
        return parse_csv(path)
    elif file_type == "pdf":
        return parse_pdf(path)
    elif file_type in ["xlsx", "xls"]:
        return parse_excel(path)
    else:
        raise ValueError(f"Unsupported file type: {file_type}")


def parse_csv(path: Path) -> dict:
    """
    Reads a CSV file and extracts its contents into a structured dictionary.
    Ideal for bank statements and transaction exports.
    """
    df = pd.read_csv(path)

    # Clean column names — remove spaces, lowercase
    df.columns = df.columns.str.strip().str.lower().str.replace(" ", "_")

    # Drop completely empty rows
    df = df.dropna(how="all")

    return {
        "type": "csv",
        "rows": len(df),
        "columns": list(df.columns),
        "data": df.head(200).to_dict(orient="records"),  # cap at 200 rows for AI
        "summary": {
            "total_rows": len(df),
            "numeric_columns": list(df.select_dtypes(include="number").columns),
        }
    }


def parse_pdf(path: Path) -> dict:
    """
    Reads a PDF and extracts both text and tables.
    Ideal for invoices and financial statements.
    """
    text_content = []
    tables = []

    with pdfplumber.open(path) as pdf:
        for i, page in enumerate(pdf.pages):
            # Extract text
            text = page.extract_text()
            if text:
                text_content.append({
                    "page": i + 1,
                    "text": text.strip()
                })

            # Extract tables
            page_tables = page.extract_tables()
            for table in page_tables:
                if table:
                    # Convert to list of dicts using first row as headers
                    headers = [str(h).strip() if h else f"col_{i}"
                               for i, h in enumerate(table[0])]
                    rows = []
                    for row in table[1:]:
                        rows.append({
                            headers[i]: str(cell).strip() if cell else ""
                            for i, cell in enumerate(row)
                        })
                    tables.append(rows)

    return {
        "type": "pdf",
        "pages": len(text_content),
        "text_content": text_content,
        "tables": tables,
        "summary": {
            "total_pages": len(text_content),
            "tables_found": len(tables),
        }
    }


def parse_excel(path: Path) -> dict:
    """
    Reads an Excel file and extracts all sheets.
    Ideal for payroll exports and accounting data.
    """
    workbook = openpyxl.load_workbook(path, data_only=True)
    sheets = {}

    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        rows = []

        for row in sheet.iter_rows(values_only=True):
            # Skip completely empty rows
            if any(cell is not None for cell in row):
                rows.append(list(row))

        if rows:
            # Use first row as headers
            headers = [str(h).strip() if h else f"col_{i}"
                       for i, h in enumerate(rows[0])]
            data = []
            for row in rows[1:]:
                data.append({
                    headers[i]: row[i]
                    for i in range(min(len(headers), len(row)))
                })
            sheets[sheet_name] = {
                "headers": headers,
                "rows": len(data),
                "data": data[:200]  # cap at 200 rows for AI
            }

    return {
        "type": "excel",
        "sheets": list(sheets.keys()),
        "data": sheets,
        "summary": {
            "total_sheets": len(sheets),
        }
    }